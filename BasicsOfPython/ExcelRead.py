import openpyxl as openpyxl
class ExcelRead():

    dict ={}
    def ReadData(self,SheetName):

        filepath = "C:\\Users\\kumar\\PycharmProjects\\FITACorePythonAPRBatch\\Input\\StudentInfo.xlsx"
        # open the excel file
        workbook = openpyxl.load_workbook(filepath)
        # move to the specific sheet
        Sheet = workbook[SheetName]
        # go to the specific row and column
        column = Sheet.cell(row=2,column=1).value
        #print(column)
        # to get a Max used row
        Total_Rows = Sheet.max_row
        # to get a max column
        Total_column = Sheet.max_column
        # to navigate in to all the rows
        for eachrow in range(2,Total_Rows+1):
            #to navigate in to all the columns
            for eachcolumn in range(1,Total_column+1):
                #get the value from each row and column
                eachCell = Sheet.cell(row=eachrow, column=eachcolumn).value
                #print(eachCell)
                #key =Sheet.cell(row=1, column=eachcolumn).value +str(eachrow)
                #key = Sheet.cell(row=1, column=eachcolumn).value + Sheet.cell(row=eachrow, column=1).value
                key = Sheet.cell(row=1, column=eachcolumn).value + str(eachrow)

                self.dict[key] = Sheet.cell(row=eachrow, column=eachcolumn).value
        print("*********Dictionalty values**********")
        print( self.dict )
        #for eachvalue in self.dict.values():
            #print(eachvalue)
        return self.dict

        #print(self.dict["FeesPaidFT2"])

    def Excel_Write(self):
        writefilepath = "C:\\Users\\kumar\\PycharmProjects\\FITACorePythonAPRBatch\\Input\\output.xlsx"
        # open the excel file
        writeworkbook = openpyxl.Workbook()
        # move to the specific sheet
        sheet = writeworkbook.active
        sheet.title = "output"
        # go to the specific row and column
        #c= sheet.cell(row = 1, column = 1)
#
        sheet.cell(row=1,column=1).value = "Sathish"
        sheet.cell(row=1, column=2).value = "kumar"
        sheet.cell(row=1, column=3).value = "R"
        sheet.cell(row=2, column=1).value = "raja"
        sheet.cell(row=2, column=2).value = "kumar"
        sheet.cell(row=2, column=3).value = "M"

        writeworkbook.save(writefilepath)
        writeworkbook.close()

    def readandwrite(self,SheetName):
        filepath = "C:\\Users\\kumar\\PycharmProjects\\FITACorePythonAPRBatch\\Input\\StudentInfo.xlsx"
        writefilepath = "C:\\Users\\kumar\\PycharmProjects\\FITACorePythonAPRBatch\\Input\\output.xlsx"

        #
        # open the excel file
        workbook = openpyxl.load_workbook(filepath)
        # open the excel file
        writeworkbook = openpyxl.Workbook()

        # move to the specific sheet
        Sheet = workbook[SheetName]
        # create asheet
        writesheet = writeworkbook.active
        writesheet.title = "output"
        # to get a Max used row
        Total_Rows = Sheet.max_row
        # to get a max column
        Total_column = Sheet.max_column
        # to navigate in to all the rows
        for eachrow in range(1, Total_Rows + 1):
            # to navigate in to all the columns
            for eachcolumn in range(1, Total_column + 1):
                # get the value from each row and column
                eachCell = Sheet.cell(row=eachrow, column=eachcolumn).value
                writesheet.cell(row=eachrow, column=eachcolumn).value = eachCell

        writeworkbook.save(writefilepath)
        writeworkbook.close()
        print("done")

obj = ExcelRead()
#obj.readandwrite("Student")
obj.ReadData("Student")